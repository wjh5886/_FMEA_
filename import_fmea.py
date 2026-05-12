"""
FMEA Excel → Supabase 임포트 스크립트
- GN7_FL, LQ2, TK1: SW FMEA / SW_FMEA 시트
- JG1, ME1: .xlsx 오픈 실패 시 xlrd 시도
"""

import openpyxl
import xlrd
import requests
import json
import os
import sys
import urllib3
from pathlib import Path

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

SUPABASE_URL = "https://itzgdbeiyvodhfhmvrfw.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Iml0emdkYmVpeXZvZGhmaG12cmZ3Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzY3NDE2MzcsImV4cCI6MjA5MjMxNzYzN30.iqzQr-3Lqf1O4UHFe9euTLyeIyBXreLPoSzUdtEaNP8"

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

VALID_MODES = {'MORE', 'LESS', 'CORRUPT', 'EARLY', 'LATE', 'STUCK', 'ERRATIC', 'N/A'}

FMEA_DIR = Path("E:/claude/FMEA/SBW_FMEA")

# 프로젝트 정의: (폴더명, 차량모델, 엑셀파일, 시트명우선순위)
PROJECTS = [
    ("GN7_FL", "GN7 FL", "GN7FL_SBW-Software FMEA_work4.xlsx", ["SW FMEA"]),
    ("LQ2",    "LQ2",    "LQ2_SBW-Software FMEA.xlsx",         ["SW_FMEA", "SW FMEA"]),
    ("TK1",    "TK1",    "TK1_SBW-Software FMEA.xlsx",         ["SW_FMEA", "SW FMEA"]),
    ("JG1",    "JG1",    "JG1_SBW-Software FMEA.xlsx",         ["SW_FMEA", "SW FMEA"]),
    ("ME1",    "ME1",    "ME1_SBW-Software FMEA.xlsx",         ["SW_FMEA", "SW FMEA"]),
]


def sb_get(table, params=None):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/{table}", headers=HEADERS, params=params, verify=False)
    r.raise_for_status()
    return r.json()


def sb_post(table, data):
    r = requests.post(f"{SUPABASE_URL}/rest/v1/{table}", headers=HEADERS, json=data, verify=False)
    r.raise_for_status()
    return r.json()


def sb_post_batch(table, rows, batch=200):
    inserted = 0
    for i in range(0, len(rows), batch):
        chunk = rows[i:i+batch]
        r = requests.post(f"{SUPABASE_URL}/rest/v1/{table}", headers=HEADERS, json=chunk, verify=False)
        if r.status_code in (200, 201):
            inserted += len(chunk)
        else:
            print(f"  !! 배치 오류: {r.status_code} {r.text[:200]}")
    return inserted


def get_or_create_project(name, vehicle_model):
    existing = sb_get("projects", {"name": f"eq.{name}", "select": "id,name"})
    if existing:
        print(f"  → 기존 프로젝트 사용: {name} ({existing[0]['id']})")
        return existing[0]["id"]
    created = sb_post("projects", {"name": name, "vehicle_model": vehicle_model, "description": f"SBW FMEA - {name}"})
    if isinstance(created, list):
        pid = created[0]["id"]
    else:
        pid = created["id"]
    print(f"  → 새 프로젝트 생성: {name} ({pid})")
    return pid


def ensure_sw_units(unit_names, project_id):
    existing = sb_get("sw_units", {"project_id": f"eq.{project_id}", "select": "id,name"})
    unit_map = {u["name"]: u["id"] for u in existing}
    for name in unit_names:
        if name and name not in unit_map:
            r = sb_post("sw_units", {"project_id": project_id, "name": name})
            uid = (r[0] if isinstance(r, list) else r)["id"]
            unit_map[name] = uid
    return unit_map


def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    # 줄바꿈 정리
    s = s.replace('\n', ' ').replace('\r', '')
    return s if s else None


def parse_int(val):
    try:
        v = int(float(str(val)))
        return v if 1 <= v <= 10 else None
    except Exception:
        return None


def detect_format(header_row):
    """헤더 행에서 포맷 감지. GN7_FL은 col7=Potential Causes, 나머지는 Effect on Module"""
    row = [str(v or '').strip() for v in header_row]
    # GN7_FL 포맷: col7에 'Causes'
    if len(row) > 7 and 'Causes' in str(row[7]):
        return 'GN7'
    return 'STANDARD'


def parse_sheet_rows(ws):
    """
    시트에서 FMEA 데이터 행을 파싱한다.
    헤더는 12행(index 11), 데이터는 14행(index 13)부터.
    merged cell 처리: 이전 행의 값을 유지.
    """
    all_rows = list(ws.iter_rows(values_only=True))

    # 헤더 행 찾기 (No, SW가 있는 행)
    header_idx = None
    for i, row in enumerate(all_rows):
        row_str = [str(v or '') for v in row]
        if 'No' in row_str and any('SW' in s for s in row_str):
            header_idx = i
            break

    if header_idx is None:
        print("  !! 헤더 행을 찾지 못함")
        return [], 'STANDARD'

    header = all_rows[header_idx]
    fmt = detect_format(header)

    # 데이터 시작: 헤더 다음 행 (설명 행) 이후
    data_start = header_idx + 2  # 설명 행 1개 건너뜀

    # 컬럼 인덱스 결정
    if fmt == 'GN7':
        col = dict(no=0, sw_unit=1, category=2, variable=3, vtype=4,
                   mode=5, detail=6, cause=7, effect_mod=8, effect_sys=9,
                   effect_sg=10, S=11, preventive=12, O=13, detection_act=14,
                   D=16, RPN=19)
    else:
        col = dict(no=0, sw_unit=1, category=2, variable=3, vtype=4,
                   mode=5, detail=6, effect_mod=7, effect_sys=8,
                   effect_sg=9, S=10, preventive=11, O=12, detection_act=13,
                   D=15, RPN=16, cm_req=17, countermeasure=18)

    rows_out = []
    prev = {}  # 이전 행의 값 캐시 (merged cell 처리)

    for row in all_rows[data_start:]:
        if all(v is None for v in row):
            continue

        def g(key, fallback_prev=True):
            idx = col.get(key)
            if idx is None:
                return None
            val = row[idx] if idx < len(row) else None
            if val is not None:
                prev[key] = val
                return val
            return prev.get(key) if fallback_prev else None

        # No 값 기준으로 새 항목 시작
        no_val = row[col['no']] if col['no'] < len(row) else None

        sw_unit = g('sw_unit')
        variable = g('variable')
        mode_raw = g('mode', fallback_prev=False)  # 실제 셀에 있는 값만

        # 최소한 mode 또는 variable이 있어야 유효한 행
        if not variable and not mode_raw:
            continue

        mode = str(mode_raw).strip() if mode_raw else None
        if mode and mode.upper() in VALID_MODES:
            mode = mode.upper()
        else:
            mode = None

        item = {
            "item_no": str(int(float(str(no_val)))) if no_val not in (None, '') else None,
            "sw_unit_name": clean(sw_unit),
            "category": clean(g('category')) if clean(g('category')) in ('External', 'Internal') else None,
            "variable_name": clean(variable) or '',
            "variable_type": clean(g('vtype')),
            "failure_mode": mode,
            "failure_detail": clean(g('detail', False)),
            "effect_module": clean(g('effect_mod', False)),
            "effect_system": clean(g('effect_sys', False)),
            "effect_safety_goal": clean(g('effect_sg', False)),
            "severity": parse_int(g('S', False)),
            "occurrence": parse_int(g('O', False)),
            "detection": parse_int(g('D', False)),
            "preventive_action": clean(g('preventive', False)),
            "detection_action": clean(g('detection_act', False)),
            "cm_required": None,
            "countermeasure": None,
        }

        if 'cm_req' in col:
            cm = row[col['cm_req']] if col['cm_req'] < len(row) else None
            item['cm_required'] = True if str(cm or '').strip().upper() == 'YES' else (False if str(cm or '').strip().upper() == 'NO' else None)
            item['countermeasure'] = clean(row[col['countermeasure']] if col['countermeasure'] < len(row) else None)

        if item['variable_name']:
            rows_out.append(item)

    return rows_out, fmt


class XlsWorksheet:
    """xlrd 시트를 openpyxl-like 인터페이스로 래핑"""
    def __init__(self, sheet):
        self._sheet = sheet

    def iter_rows(self, values_only=True):
        for rx in range(self._sheet.nrows):
            yield tuple(
                self._sheet.cell(rx, cx).value
                for cx in range(self._sheet.ncols)
            )


class XlsWorkbook:
    def __init__(self, wb):
        self._wb = wb

    @property
    def sheetnames(self):
        return self._wb.sheet_names()

    def __getitem__(self, name):
        return XlsWorksheet(self._wb.sheet_by_name(name))

    def close(self):
        self._wb.release_resources()


def load_workbook_safe(fpath):
    """openpyxl 시도, 실패시 xlrd 시도"""
    with open(fpath, 'rb') as f:
        magic = f.read(2)
    if magic == b'PK':
        try:
            return openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"  !! openpyxl 실패: {e}")
            return None
    else:
        try:
            wb = xlrd.open_workbook(fpath)
            print(f"  (XLS binary 포맷으로 읽기)")
            return XlsWorkbook(wb)
        except Exception as e:
            print(f"  !! xlrd 실패: {e}")
            return None


def import_project(folder, model, xlsx_name, sheet_priorities):
    fpath = FMEA_DIR / folder / xlsx_name
    if not fpath.exists():
        print(f"  !! 파일 없음: {fpath}")
        return

    print(f"\n{'='*60}")
    print(f"프로젝트: {folder} ({model})")
    print(f"파일: {xlsx_name}")

    wb = load_workbook_safe(str(fpath))
    if wb is None:
        print(f"  !! 파일을 열 수 없음 - 건너뜀")
        return

    # 시트 선택
    ws = None
    for sname in sheet_priorities:
        if sname in wb.sheetnames:
            ws = wb[sname]
            print(f"  시트: {sname}")
            break

    if ws is None:
        print(f"  !! 해당 시트 없음 (사용 가능: {wb.sheetnames})")
        wb.close()
        return

    rows, fmt = parse_sheet_rows(ws)
    wb.close()

    print(f"  파싱된 행 수: {len(rows)} (포맷: {fmt})")
    if not rows:
        print("  !! 데이터 없음")
        return

    # 프로젝트 생성/조회
    project_id = get_or_create_project(folder, model)

    # SW Unit 목록 추출 및 생성
    unit_names = list({r['sw_unit_name'] for r in rows if r['sw_unit_name']})
    unit_map = ensure_sw_units(unit_names, project_id)
    print(f"  SW Units: {len(unit_map)}개")

    # fmea_items 준비
    items = []
    for r in rows:
        items.append({
            "project_id": project_id,
            "sw_unit_id": unit_map.get(r['sw_unit_name']),
            "item_no": r['item_no'],
            "category": r['category'],
            "variable_name": r['variable_name'],
            "variable_type": r['variable_type'],
            "failure_mode": r['failure_mode'],
            "failure_detail": r['failure_detail'],
            "effect_module": r['effect_module'],
            "effect_system": r['effect_system'],
            "effect_safety_goal": r['effect_safety_goal'],
            "severity": r['severity'],
            "occurrence": r['occurrence'],
            "detection": r['detection'],
            "preventive_action": r['preventive_action'],
            "detection_action": r['detection_action'],
            "cm_required": r['cm_required'],
            "countermeasure": r['countermeasure'],
            "status": "draft",
        })

    inserted = sb_post_batch("fmea_items", items)
    print(f"  >> 삽입 완료: {inserted}/{len(items)}개")


def main():
    print("FMEA Excel 임포트 시작")
    print(f"Supabase: {SUPABASE_URL}")

    # 연결 테스트
    try:
        r = sb_get("projects", {"select": "id", "limit": "1"})
        print(f"Supabase 연결 OK\n")
    except Exception as e:
        print(f"Supabase 연결 실패: {e}")
        sys.exit(1)

    for folder, model, xlsx, sheets in PROJECTS[3:]:
        import_project(folder, model, xlsx, sheets)

    print("\n\n임포트 완료!")


if __name__ == "__main__":
    main()
