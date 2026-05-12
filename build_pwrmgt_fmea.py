#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CstAp_PwrMGT Software FMEA - 처음부터 완성 생성기
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

NA = None

def NA_ROW(mode, reason):
    return (NA, NA, NA, NA, mode, reason, NA, NA, NA, NA,
            NA, NA, NA, NA, NA, NA, NA, NA)

# ─────────────────────────────────────────────────────────────────────────────
# Variable definitions
# ─────────────────────────────────────────────────────────────────────────────
VAR1 = ("TrmnlCtrlGrpStaBDCEV\n"
        "(BDCEV_OFF : 0u\nBDCEV_POWERON : 1u\nBDCEV_CRANKING : 2u\nBDCEV_READY : 3u)\n"
        "[CAN] SMK_TrmnlCtrlGrpStaBDCEV\n"
        "  MSG: BDC_FD_SMK_03_200ms (0x402)\n"
        "  Bit: 96|3bit, Sender: CGW_CCU")
VAR1_TYPE = "uint8\nRange: 0~3\n(CAN 3bit → OOR: 4~7)"
VAR1_SM = "[SM09] IGN Signal Redundancy - High\n[SM08] MCU Safety Mechanism(Memory) - High"

VAR2 = "ECUSta\n(EXTER_ECU_WAKEUP : 1u\nEXTER_ECU_DOOR_WAKEUP : 2u\nEXTER_ECU_STANDBY : 3u\nEXTER_ECU_SLEEP : 4u)"
VAR2_TYPE = "uint8\nRange: 1~4"
VAR2_SM = "[SM08] MCU Safety Mechanism(Memory) - High"

VAR3 = "BatVolt\n(0 ~ 4095u)"
VAR3_TYPE = "uint16\nRange: 0~4095"
VAR3_SM = "[SM08] MCU Safety Mechanism(Memory) - High"

VAR4 = "IgnVolt\n(0 ~ 4095u)"
VAR4_TYPE = "uint16\nRange: 0~4095"
VAR4_SM = "[SM08] MCU Safety Mechanism(Memory) - High\n[SM09] IGN Signal Redundancy - High"

VAR5 = "SbcFlt\n(Normal : 0u\nSBC Fault : 1u)"
VAR5_TYPE = "uint8\nRange: 0~1"
VAR5_SM = "[SM08] MCU Safety Mechanism(Memory) - High"

VAR6 = "Ldo2OnVolt\n(0 ~ 4095u)"
VAR6_TYPE = "uint16\nRange: 0~4095"
VAR6_SM = "[SM08] MCU Safety Mechanism(Memory) - High\n[SM10] Power Supply#2 Voltage Monitoring - High"

TEST = "1. 단위 테스트/LBIST"
DR = "Design Review"

# ─────────────────────────────────────────────────────────────────────────────
# Row data
# Each tuple: (sw_unit, category, var_name, var_type,
#              fm_mode, fm_detail,
#              effect_module, cause, effect_system, sg,
#              S, prev_action, O, safety_mechanism, test_sm,
#              D_sm, D_test, countermeasure)
# N/A rows: all None except fm_mode and fm_detail
# ─────────────────────────────────────────────────────────────────────────────
rows_data = [
    # ═══════════════════════════════════════════════════════════════════════════
    # Variable 1: TrmnlCtrlGrpStaBDCEV
    # ═══════════════════════════════════════════════════════════════════════════
    ("CstAp_PwrMGT", "External", VAR1, VAR1_TYPE,
     "MORE", "From : BDCEV_OFF\nTo : BDCEV_POWERON",
     "비정상 BDCEV_POWERON 인식\n(PowerOnSta ON 오판정)",
     "HW Random Failure\nBDC_FD_SMK_03_200ms CAN 신호 오류\n(CGW_CCU Tx 오류)",
     "Normal 전환 오인식\n  - B/L On, DTC 발생 안함",
     "X", 1, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "MORE", "From : BDCEV_OFF\nTo : BDCEV_CRANKING",
     "비정상 BDCEV_CRANKING 인식",
     "HW Random Failure\nBDC_FD_SMK_03_200ms CAN 신호 오류\n(CGW_CCU Tx 오류)",
     "DTC 발생 불가",
     "X", 1, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "MORE", "From : BDCEV_OFF\nTo : BDCEV_READY",
     "비정상 BDCEV_READY 인식 (최고 상태 오전환)",
     "HW Random Failure\nCAN 신호 오류",
     "차량제어 오동작 가능성\n  - 해당, 이전 DTC 발생 안함",
     "X", 1, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "MORE", "From : BDCEV_POWERON\nTo : BDCEV_CRANKING",
     "비정상 BDCEV_CRANKING 인식",
     "HW Random Failure\nCAN 신호 오류",
     "DTC 발생 불가",
     "X", 1, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "MORE", "From : BDCEV_POWERON\nTo : BDCEV_READY",
     "비정상 BDCEV_READY 인식",
     "HW Random Failure\nCAN 신호 오류",
     "차량제어 오동작 가능성\n  - 해당, 이전 DTC 발생 안함",
     "X", 1, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "MORE", "From : BDCEV_CRANKING\nTo : BDCEV_READY",
     "비정상 BDCEV_READY 인식",
     "HW Random Failure\nCAN 신호 오류",
     "차량제어 오동작 가능성\n  - 해당, 이전 DTC 발생 안함",
     "X", 1, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : BDCEV_READY\nTo : BDCEV_CRANKING",
     "비정상 BDCEV_CRANKING 인식",
     "HW Random Failure\nCAN 신호 오류",
     "DTC 발생 불가",
     "X", 1, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : BDCEV_READY\nTo : BDCEV_POWERON",
     "비정상 BDCEV_POWERON 인식",
     "HW Random Failure\nCAN 신호 오류",
     "Normal 전환 오인식\n  - B/L On, DTC 발생 안함",
     "X", 1, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : BDCEV_READY\nTo : BDCEV_OFF",
     "비정상 BDCEV_OFF 인식",
     "HW Random Failure\nCAN 신호 오류",
     "DTC 발생 불가",
     "X", 1, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : BDCEV_CRANKING\nTo : BDCEV_POWERON",
     "비정상 BDCEV_POWERON 인식",
     "HW Random Failure\nCAN 신호 오류",
     "Normal 전환 오인식\n  - B/L On, DTC 발생 안함",
     "X", 1, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : BDCEV_CRANKING\nTo : BDCEV_OFF",
     "비정상 BDCEV_OFF 인식",
     "HW Random Failure\nCAN 신호 오류",
     "DTC 발생 불가",
     "X", 1, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : BDCEV_POWERON\nTo : BDCEV_OFF",
     "비정상 BDCEV_OFF 인식",
     "HW Random Failure\nCAN 신호 오류",
     "DTC 발생 불가",
     "X", 1, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    NA_ROW("REVERSE", "- Boolean, Signed Data 형 아니므로 해당 없음"),

    # CAN DB: SMK_TrmnlCtrlGrpStaBDCEV is 3bit → OOR = value 4~7
    (NA, NA, NA, NA, "CORRUPT", "From : BDCEV_OFF (0)\nTo : Out of Range (4~7)\n[CAN 3bit 신호, 유효값 0~3]",
     "Power Off 오인식 (CAN 3bit OOR 값 수신)",
     "BDC_FD_SMK_03_200ms CAN 비트 오류\n(SGW bit corruption / HW Random Failure)",
     "BDCEV_READY, BDCEV_POWERON 사용 불가\n  - 전체 기능 사용 불가",
     "SG01, SG02, SG06", 9, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "CORRUPT", "From : BDCEV_POWERON (1)\nTo : Out of Range (4~7)\n[CAN 3bit 신호, 유효값 0~3]",
     "Power Off 오인식 (CAN 3bit OOR 값 수신)",
     "BDC_FD_SMK_03_200ms CAN 비트 오류\n(SGW bit corruption / HW Random Failure)",
     "BDCEV_READY, BDCEV_POWERON 사용 불가\n  - 전체 기능 사용 불가",
     "SG01, SG02, SG06", 9, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "CORRUPT", "From : BDCEV_CRANKING (2)\nTo : Out of Range (4~7)\n[CAN 3bit 신호, 유효값 0~3]",
     "Power Off 오인식 (CAN 3bit OOR 값 수신)",
     "BDC_FD_SMK_03_200ms CAN 비트 오류\n(SGW bit corruption / HW Random Failure)",
     "BDCEV_READY, BDCEV_POWERON 사용 불가\n  - 전체 기능 사용 불가",
     "SG01, SG02, SG06", 9, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "CORRUPT", "From : BDCEV_READY (3)\nTo : Out of Range (4~7)\n[CAN 3bit 신호, 유효값 0~3]",
     "Power Off 오인식 (CAN 3bit OOR 값 수신)",
     "BDC_FD_SMK_03_200ms CAN 비트 오류\n(SGW bit corruption / HW Random Failure)",
     "BDCEV_READY, BDCEV_POWERON 사용 불가\n  - 전체 기능 사용 불가",
     "SG01, SG02, SG06", 9, DR, 4, VAR1_SM, TEST, 1, 4, "N"),

    NA_ROW("NO", "- 열거형 신호이므로 해당 없음"),
    NA_ROW("AS WELL AS", "- 열거형 신호이므로 해당 없음"),
    NA_ROW("PART OF", "- 열거형 신호이므로 해당 없음"),
    # CAN DB: BDC_FD_SMK_03_200ms → 200ms 주기 → EARLY/LATE 의미 있음
    (NA, NA, NA, NA, "EARLY",
     "CAN 신호 조기 수신\n(BDC_FD_SMK_03_200ms 200ms 주기 전 수신)",
     "상태 전환 조기 감지 → PowerOnSta 조기 변경",
     "CGW_CCU Tx 타이밍 이상\n(200ms 이전 전송)",
     "가용성 영향 (조기 전원 전환 판정)",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),
    (NA, NA, NA, NA, "LATE",
     "CAN 신호 지연 수신\n(BDC_FD_SMK_03_200ms 200ms 초과 지연)",
     "상태 전환 지연 감지 → PowerOnSta 지연 변경",
     "CGW_CCU Tx 지연 / CAN Bus 부하 과다\n(200ms 초과)",
     "가용성 영향\n(SM05 CAN Monitoring 감지: 200ms Timeout)",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),

    # ═══════════════════════════════════════════════════════════════════════════
    # Variable 2: ECUSta
    # ═══════════════════════════════════════════════════════════════════════════
    ("CstAp_PwrMGT", "External", VAR2, VAR2_TYPE,
     "MORE", "From : EXTER_ECU_WAKEUP\nTo : EXTER_ECU_DOOR_WAKEUP",
     "Door Wakeup 상태로 오인식\n(Ldo2FltChkEn 비활성화)",
     "HW Random Failure",
     "가용성 영향\n(Ldo2 Fault 검출 불가)",
     "X", 1, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "MORE", "From : EXTER_ECU_WAKEUP\nTo : EXTER_ECU_STANDBY",
     "Standby 상태로 오인식\n(Ldo2FltChkEn 비활성화)",
     "HW Random Failure",
     "Idt 기능 All Off\n(Wakeup 중 비의도적 Standby 전환)",
     "X", 8, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "MORE", "From : EXTER_ECU_WAKEUP\nTo : EXTER_ECU_SLEEP",
     "Sleep 상태로 오인식\n(Ldo2FltChkEn 비활성화)",
     "HW Random Failure",
     "Idt 기능 All Off\n(Wakeup 중 비의도적 Sleep 전환)",
     "X", 8, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "MORE", "From : EXTER_ECU_DOOR_WAKEUP\nTo : EXTER_ECU_STANDBY",
     "Standby 상태로 오인식",
     "HW Random Failure",
     "Idt 기능 All Off\n(Door Wakeup 중 비의도적 Standby 전환)",
     "X", 8, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "MORE", "From : EXTER_ECU_DOOR_WAKEUP\nTo : EXTER_ECU_SLEEP",
     "Sleep 상태로 오인식",
     "HW Random Failure",
     "Idt 기능 All Off\n(Door Wakeup 중 비의도적 Sleep 전환)",
     "X", 8, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "MORE", "From : EXTER_ECU_STANDBY\nTo : EXTER_ECU_SLEEP",
     "Sleep 상태로 오인식",
     "HW Random Failure",
     "가용성 영향\n(Standby→Sleep 오전환, 복구 지연)",
     "X", 1, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : EXTER_ECU_SLEEP\nTo : EXTER_ECU_STANDBY",
     "Standby 상태로 오인식",
     "HW Random Failure",
     "가용성 영향\n(Sleep→Standby 오전환)",
     "X", 1, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : EXTER_ECU_SLEEP\nTo : EXTER_ECU_DOOR_WAKEUP",
     "Door Wakeup 상태로 오인식\n(불필요한 Ldo2 점검 시작)",
     "HW Random Failure",
     "비의도적 Door Wakeup 시퀀스 동작",
     "X", 6, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : EXTER_ECU_SLEEP\nTo : EXTER_ECU_WAKEUP",
     "Wakeup 상태로 오인식\n(Ldo2FltChkEn 활성화됨)",
     "HW Random Failure",
     "비의도적 Wakeup 시퀀스 동작",
     "X", 6, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : EXTER_ECU_STANDBY\nTo : EXTER_ECU_DOOR_WAKEUP",
     "Door Wakeup 상태로 오인식",
     "HW Random Failure",
     "비의도적 Door Wakeup 시퀀스 동작",
     "X", 6, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : EXTER_ECU_STANDBY\nTo : EXTER_ECU_WAKEUP",
     "Wakeup 상태로 오인식\n(Ldo2FltChkEn 활성화됨)",
     "HW Random Failure",
     "비의도적 Wakeup 시퀀스 동작",
     "X", 6, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : EXTER_ECU_DOOR_WAKEUP\nTo : EXTER_ECU_WAKEUP",
     "Wakeup 상태로 오인식\n(Ldo2FltChkEn 활성화됨)",
     "HW Random Failure",
     "가용성 영향\n(Door Wakeup→Wakeup 오전환)",
     "X", 1, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    NA_ROW("REVERSE", "- Boolean, Signed Data 형 아니므로 해당 없음"),

    (NA, NA, NA, NA, "CORRUPT", "From : EXTER_ECU_WAKEUP\nTo : Out of Range",
     "범위 이탈 값으로 상태 오판정\n(기본값 STANDBY 적용됨)",
     "HW Random Failure",
     "Ldo2FltSta 측정 불가 및 목표위치 불가",
     "SG06", 9, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "CORRUPT", "From : EXTER_ECU_DOOR_WAKEUP\nTo : Out of Range",
     "범위 이탈 값으로 상태 오판정\n(기본값 STANDBY 적용됨)",
     "HW Random Failure",
     "Ldo2FltSta 측정 불가 및 목표위치 불가",
     "SG06", 9, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "CORRUPT", "From : EXTER_ECU_STANDBY\nTo : Out of Range",
     "범위 이탈 값으로 상태 오판정\n(기본값 STANDBY 적용됨)",
     "HW Random Failure",
     "Ldo2FltSta 측정 불가 및 목표위치 불가",
     "SG06", 9, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "CORRUPT", "From : EXTER_ECU_SLEEP\nTo : Out of Range",
     "범위 이탈 값으로 상태 오판정\n(기본값 STANDBY 적용됨)",
     "HW Random Failure",
     "Ldo2FltSta 측정 불가 및 목표위치 불가",
     "SG06", 9, DR, 4, VAR2_SM, TEST, 3, 4, "N"),

    NA_ROW("NO", "- 해당 없음 (Range 1~4, 0 미정의)"),
    (NA, NA, NA, NA, "AS WELL AS", "신호 중복 입력 발생",
     "비정상 ECUSta 중복 인식", "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),
    NA_ROW("PART OF", "- 해당 없음"),
    (NA, NA, NA, NA, "EARLY", "신호 입력 빠름 (ECU 모드 전환 예상보다 이른 수신)",
     "Ldo2 점검 Early 활성화", "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),
    (NA, NA, NA, NA, "LATE", "신호 입력 늦음 (ECU 모드 전환 지연)",
     "Ldo2 점검 Late 활성화", "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),

    # ═══════════════════════════════════════════════════════════════════════════
    # Variable 3: BatVolt
    # ═══════════════════════════════════════════════════════════════════════════
    ("CstAp_PwrMGT", "External", VAR3, VAR3_TYPE,
     "MORE", "From : Low Bat (8.5V Under)\nTo : Normal (9V~16V)",
     "Normal Battery 상태로 오인식\n(BatUnderSta OFF 오판정)",
     "HW Random Failure",
     "가용성 영향\n(저전압에서 정상 동작 시도)",
     "X", 1, DR, 4, VAR3_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "MORE", "From : Low Bat (8.5V Under)\nTo : High Bat (16.5V Over)",
     "High Battery 상태로 오인식\n(BatOverSta ON 오판정)",
     "HW Random Failure",
     "차량 이동 중이라면\n과전압 상태로 판정, 이동 DTC 발생",
     "X", 1, DR, 4, VAR3_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "MORE", "From : Normal (9V~16V)\nTo : High Bat (16.5V Over)",
     "High Battery 상태로 오인식\n(BatOverSta ON 오판정)",
     "HW Random Failure",
     "차량 이동 중이라면\n과전압 상태로 판정, 이동 DTC 발생",
     "X", 1, DR, 4, VAR3_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : High Bat (16.5V Over)\nTo : Normal (9V~16V)",
     "Normal Battery 상태로 오인식\n(BatOverSta OFF 오판정)",
     "HW Random Failure",
     "가용성 영향\n(과전압 복구 DTC 미발생)",
     "X", 1, DR, 4, VAR3_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : High Bat (16.5V Over)\nTo : Low Bat (8.5V Under)",
     "Low Battery 상태로 오인식\n(BatUnderSta ON 오판정)",
     "HW Random Failure",
     "차량 이동 중이라면\n저전압 상태로 판정, 이동 DTC 발생",
     "X", 1, DR, 4, VAR3_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : Normal (9V~16V)\nTo : Low Bat (8.5V Under)",
     "Low Battery 상태로 오인식\n(BatUnderSta ON 오판정)",
     "HW Random Failure",
     "차량 이동 중이라면\n저전압 상태로 판정, 이동 DTC 발생",
     "X", 1, DR, 4, VAR3_SM, TEST, 3, 4, "N"),

    NA_ROW("REVERSE", "- Boolean, Signed Data 형 아니므로 해당 없음"),
    NA_ROW("CORRUPT", "- Range 전체를 사용하는 값이므로 해당 없음"),
    NA_ROW("NO", "- 해당 없음 (0 = ADC 최소값, 정상 범위 내)"),
    (NA, NA, NA, NA, "AS WELL AS", "신호 중복 입력 발생",
     "비정상 BatVolt 중복 인식", "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),
    NA_ROW("PART OF", "- 해당 없음"),
    (NA, NA, NA, NA, "EARLY", "배터리 전압 ADC 샘플 빠름 (debounce 조기 만족)",
     "BatStbSta 조기 판정", "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),
    (NA, NA, NA, NA, "LATE", "배터리 전압 ADC 샘플 늦음 (debounce 지연)",
     "BatStbSta 판정 지연", "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),

    # ═══════════════════════════════════════════════════════════════════════════
    # Variable 4: IgnVolt
    # ═══════════════════════════════════════════════════════════════════════════
    ("CstAp_PwrMGT", "External", VAR4, VAR4_TYPE,
     "MORE", "From : 4V Under (IGN OFF)\nTo : 7V Over (IGN ON)",
     "IGN ON 상태로 오인식\n(VIgnOnFlag=ON 오판정 → PowerOnSta=ON)",
     "HW Random Failure",
     "IGN OFF 상태에서 MCU 전체 동작 가능성\n  - 모든 기능 활성화로 안전 위험",
     "SG01, SG02, SG03, SG04, SG05, SG06", 10, DR, 4, VAR4_SM, TEST, 1, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : 7V Over (IGN ON)\nTo : 4V Under (IGN OFF)",
     "IGN OFF 상태로 오인식\n(VIgnOnFlag=OFF → PowerOnSta=OFF)",
     "HW Random Failure",
     "잘못된 IGN ON 상태에서 DTC 미발생 시\nN→다른 Mode로 전환될 수 있음",
     "SG01, SG02, SG03, SG04, SG05, SG06", 10, DR, 4, VAR4_SM, TEST, 1, 4, "N"),

    NA_ROW("REVERSE", "- Boolean, Signed Data 형 아니므로 해당 없음"),
    NA_ROW("CORRUPT", "- Range 전체를 사용하는 값이므로 해당 없음"),
    NA_ROW("NO", "- 해당 없음 (0V = IGN OFF, 정상 상태)"),
    (NA, NA, NA, NA, "AS WELL AS", "신호 중복 입력 발생",
     "비정상 IgnVolt 중복 인식", "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),
    NA_ROW("PART OF", "- 해당 없음"),
    (NA, NA, NA, NA, "EARLY", "IGN 전압 ADC 샘플 빠름 (debounce 카운터 조기 만족)",
     "VIgnOnFlag 조기 ON 판정", "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),
    (NA, NA, NA, NA, "LATE", "IGN 전압 ADC 샘플 늦음 (debounce 카운터 지연)",
     "VIgnOnFlag 지연 ON → PowerOnSta 지연", "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),

    # ═══════════════════════════════════════════════════════════════════════════
    # Variable 5: SbcFlt
    # ═══════════════════════════════════════════════════════════════════════════
    ("CstAp_PwrMGT", "External", VAR5, VAR5_TYPE,
     "MORE", "From : Normal (0)\nTo : SBC Fault (1)",
     "SBC Fault 상태로 오인식\n(SbcFltSta=ON → SysPwrSta=OFF)",
     "HW Random Failure",
     "정상 상태에서 SysPwr Off 상태로 변경\n  - 전체 SBW 기능 중지",
     "SG01, SG02, SG06", 9, DR, 4, VAR5_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : SBC Fault (1)\nTo : Normal (0)",
     "SBC Fault 숨김\n(SbcFltSta=OFF → SysPwrSta=ON 유지)",
     "HW Random Failure",
     "SBC 고장 상태에서 SysPwr On 유지\n  - 비정상 동작 지속 (모든 기능 비정상)",
     "SG01, SG02, SG03, SG04, SG05, SG06", 10, DR, 4, VAR5_SM, TEST, 3, 4, "N"),

    NA_ROW("REVERSE", "- Boolean, Signed Data 형 아니므로 해당 없음"),

    (NA, NA, NA, NA, "CORRUPT", "From : Normal (0)\nTo : Out of Range",
     "범위 이탈 값으로 SBC 상태 오판정\n(필터 카운터 비정상 동작)",
     "HW Random Failure",
     "정상 상태에서 SysPwr Off 상태로 변경\n  - 전체 SBW 기능 중지",
     "SG01, SG02, SG03, SG04, SG05, SG06", 10, DR, 4, VAR5_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "CORRUPT", "From : SBC Fault (1)\nTo : Out of Range",
     "범위 이탈 값으로 SBC Fault 숨김 가능\n(필터 카운터 비정상 동작)",
     "HW Random Failure",
     "가용성 영향\n(SBC Fault 상태 유사, SysPwr Off 지속)",
     "X", 1, DR, 4, VAR5_SM, TEST, 3, 4, "N"),

    NA_ROW("NO", "- 해당 없음 (0 = Normal 상태, 기본값)"),
    (NA, NA, NA, NA, "AS WELL AS", "신호 중복 입력 발생",
     "비정상 SbcFlt 중복 인식", "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),
    NA_ROW("PART OF", "- 해당 없음"),
    (NA, NA, NA, NA, "EARLY", "SBC Fault 신호 조기 수신 (필터 시간 전 감지)",
     "SbcFltSta 조기 ON → SysPwrSta 조기 OFF",
     "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),
    (NA, NA, NA, NA, "LATE", "SBC Fault 신호 지연 수신 (필터 카운터 도달 지연)",
     "SbcFltSta 판정 지연", "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),

    # ═══════════════════════════════════════════════════════════════════════════
    # Variable 6: Ldo2OnVolt
    # ═══════════════════════════════════════════════════════════════════════════
    ("CstAp_PwrMGT", "External", VAR6, VAR6_TYPE,
     "MORE", "From : Under Voltage (4.5V Under)\nTo : Normal (4.8V Over)",
     "LDO2 정상 상태로 오인식\n(Ldo2UnderFlag=OFF → Ldo2FltSta=OFF)",
     "HW Random Failure",
     "LDO2 고장 미감지\n  - Indicator 비정상 전압 공급 지속\n  (현재 Ldo2FltSta 출력 비활성 상태)",
     "SG04", 8, DR, 4, VAR6_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "MORE", "From : Under Voltage (4.5V Under)\nTo : Out of Range",
     "범위 이탈 값으로 LDO2 정상 오인식",
     "HW Random Failure",
     "LDO2 고장 미감지\n  (현재 Ldo2FltSta 출력 비활성 상태)",
     "SG04", 8, DR, 4, VAR6_SM, TEST, 3, 4, "N"),

    (NA, NA, NA, NA, "LESS", "From : Normal (4.8V Over)\nTo : Under Voltage (4.5V Under)",
     "LDO2 Under Voltage 오인식\n(Ldo2UnderFlag=ON → Ldo2FltSta=ON)",
     "HW Random Failure",
     "정상 LDO2 전압에서 Ldo2 Fault 오판정\n  - Indicator Off (가용성 영향)",
     "X", 1, DR, 4, VAR6_SM, TEST, 3, 4, "N"),

    NA_ROW("REVERSE", "- Boolean, Signed Data 형 아니므로 해당 없음"),

    (NA, NA, NA, NA, "CORRUPT", "From : Normal\nTo : Out of Range",
     "범위 이탈 ADC 값 수신\n(Ldo2 전압 측정 불가)",
     "HW Random Failure",
     "LDO2 Fault 오판정 또는 미판정\n  (현재 Ldo2FltSta 출력 비활성 상태)",
     "SG04", 8, DR, 4, VAR6_SM, TEST, 3, 4, "N"),

    NA_ROW("NO", "- 해당 없음 (0V = ADC 최소값, 유효 범위 내)"),
    (NA, NA, NA, NA, "AS WELL AS", "신호 중복 입력 발생",
     "비정상 Ldo2OnVolt 중복 인식", "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),
    NA_ROW("PART OF", "- 해당 없음"),
    (NA, NA, NA, NA, "EARLY", "LDO2 전압 ADC 샘플 빠름 (1s 필터 조기 만족)",
     "Ldo2FltSta 조기 ON 판정", "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),
    (NA, NA, NA, NA, "LATE", "LDO2 전압 ADC 샘플 늦음 (1s 필터 지연)",
     "Ldo2FltSta 판정 지연 → Indicator 비정상 공급 지속",
     "HW Random Failure", "가용성 영향",
     "-", NA, NA, NA, NA, NA, NA, NA, NA),
]

# ─────────────────────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────────────────────
def thin_border():
    s = Side(border_style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def wrap_center():
    return Alignment(wrap_text=True, horizontal="center", vertical="center")

def wrap_left():
    return Alignment(wrap_text=True, horizontal="left", vertical="center")

BOLD = Font(bold=True, size=10)
NORMAL = Font(size=10)

# ─────────────────────────────────────────────────────────────────────────────
# Build workbook
# ─────────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SW_FMEA"

# Column widths (A..T)
col_widths = {
    1: 6,   # No
    2: 20,  # SW Unit Name
    3: 14,  # Category
    4: 30,  # Variable name
    5: 18,  # Variable type
    6: 14,  # Failure mode
    7: 32,  # Detail
    8: 30,  # Effect on Module
    9: 30,  # Potential Cause
    10: 35, # Effect on System
    11: 22, # Effect on SG
    12: 5,  # S
    13: 22, # Preventive Action
    14: 5,  # O
    15: 42, # Safety Mechanism
    16: 28, # Test Method (SM)
    17: 5,  # D_sm
    18: 5,  # D_test
    19: 5,  # D
    20: 7,  # RPN
    21: 14, # Countermeasure
}
for col_idx, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── Title row ────────────────────────────────────────────────
ws.merge_cells("A1:U1")
title_cell = ws["A1"]
title_cell.value = "Software FMEA - CstAp_PwrMGT (JG1 SBW)"
title_cell.font = Font(bold=True, size=14)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
title_cell.fill = header_fill("1F3864")
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
ws.row_dimensions[1].height = 30

# ── Meta info ───────────────────────────────────────────────
meta = [
    ("A2", "Vehicle:", "B2", "JG1"),
    ("A3", "Item:", "B3", "SBW"),
    ("A4", "SW Unit:", "B4", "CstAp_PwrMGT"),
    ("A5", "Version:", "B5", "1.0"),
    ("D2", "Created:", "E2", "2026-04-15"),
    ("D3", "Author:", "E3", "SW Engineering Team"),
    ("D4", "Source Code:", "E4", "HKMC_JG1_SBW_R44_4.2109.00"),
]
for lbl_cell, lbl_val, val_cell, val_val in meta:
    c1 = ws[lbl_cell]
    c1.value = lbl_val
    c1.font = Font(bold=True, size=10)
    c2 = ws[val_cell]
    c2.value = val_val
    c2.font = NORMAL

# ── Severity legend ─────────────────────────────────────────
ws.merge_cells("H2:K2")
ws["H2"].value = "Severity 기준"
ws["H2"].font = Font(bold=True, size=10)
ws["H2"].alignment = wrap_center()
ws["H2"].fill = header_fill("BDD7EE")

severity_legend = [
    ("H3", "S=10", "I3", "MCU 전체 비정상 동작"),
    ("H4", "S=9",  "I4", "전체 기능 불가 (CORRUPT 수신)"),
    ("H5", "S=8",  "I5", "Idt 기능 All Off"),
    ("H6", "S=6",  "I6", "비의도적 시퀀스"),
    ("H7", "S=4",  "I7", "일반 가용성 영향"),
    ("H8", "S=1",  "I8", "가용성 영향 없음"),
]
ws.merge_cells("I3:K3")
ws.merge_cells("I4:K4")
ws.merge_cells("I5:K5")
ws.merge_cells("I6:K6")
ws.merge_cells("I7:K7")
ws.merge_cells("I8:K8")
for s_cell, s_val, d_cell, d_val in severity_legend:
    ws[s_cell].value = s_val
    ws[s_cell].font = NORMAL
    ws[d_cell].value = d_val
    ws[d_cell].font = NORMAL

# ── Column header row ───────────────────────────────────────
HDR_ROW = 10
FILL_HDR = header_fill("1F3864")
FILL_SUB = header_fill("2E75B6")
FONT_HDR = Font(bold=True, size=10, color="FFFFFF")

headers = [
    "No", "SW Unit Name", "Interface\nCategory",
    "Interface\n(Variable) name", "Interface\n(Variable) type",
    "Failure Mode\n(HAZOP)", "Detail of Failure Mode",
    "Effect on Module", "Potential Cause",
    "Effect on System", "Effect on SG\n(SG ID)",
    "S", "Preventive Action", "O",
    "Safety Mechanism", "Test Method",
    "D\n(SM)", "D\n(Test)", "D", "RPN",
    "Countermeasure\nRequired"
]
for col_i, hdr in enumerate(headers, start=1):
    cell = ws.cell(row=HDR_ROW, column=col_i, value=hdr)
    cell.font = FONT_HDR
    cell.fill = FILL_HDR
    cell.alignment = wrap_center()
    cell.border = thin_border()
ws.row_dimensions[HDR_ROW].height = 40

# ─────────────────────────────────────────────────────────────────────────────
# Data rows
# ─────────────────────────────────────────────────────────────────────────────
START_ROW = HDR_ROW + 1
ROW_NO_COUNTER = [0]

# Track merge ranges for variable name / type / sw_unit
var_start_row = None
var_end_row = None
current_var_name = None
merge_ranges = []  # [(start_row, end_row, col_list)]

# Color for N/A rows
FILL_NA = header_fill("F2F2F2")
FILL_CONTENT = header_fill("FFFFFF")
FILL_CORRUPT = header_fill("FFF2CC")
FILL_HIGH_SEV = header_fill("FCE4D6")

def row_fill(fm_mode, S):
    if S is None:
        return FILL_NA
    if fm_mode == "CORRUPT":
        return FILL_CORRUPT
    if S >= 9:
        return FILL_HIGH_SEV
    return FILL_CONTENT

# variable tracking for merge
var_block_rows = {}  # var_name -> [start_excel_row, end_excel_row]
current_var_name = None
current_var_start_excel = None

for row_i, row in enumerate(rows_data):
    excel_row = START_ROW + row_i
    (sw_unit, category, var_name, var_type,
     fm_mode, fm_detail,
     effect_module, cause, effect_system, sg,
     S, prev_action, O, sm, test_sm,
     D_sm, D_test, countermeasure) = row

    # Track var block
    if var_name is not None:
        if current_var_name != var_name:
            if current_var_name is not None:
                var_block_rows[current_var_name][1] = excel_row - 1
            current_var_name = var_name
            var_block_rows[var_name] = [excel_row, excel_row]
    # update end
    if current_var_name is not None:
        var_block_rows[current_var_name][1] = excel_row

    fill = row_fill(fm_mode, S)

    # No (col 1) — only for content rows
    if S is not None:
        ROW_NO_COUNTER[0] += 1
        ws.cell(excel_row, 1).value = ROW_NO_COUNTER[0]

    # SW Unit (col 2)
    ws.cell(excel_row, 2).value = sw_unit or ""

    # Category (col 3)
    ws.cell(excel_row, 3).value = category or ""

    # Variable name (col 4)
    ws.cell(excel_row, 4).value = var_name or ""

    # Variable type (col 5)
    ws.cell(excel_row, 5).value = var_type or ""

    # Failure Mode (col 6)
    ws.cell(excel_row, 6).value = fm_mode or ""

    # Failure Mode Detail (col 7)
    ws.cell(excel_row, 7).value = fm_detail or ""

    # Effect on Module (col 8)
    ws.cell(excel_row, 8).value = effect_module or ""

    # Cause (col 9)
    ws.cell(excel_row, 9).value = cause or ""

    # Effect on System (col 10)
    ws.cell(excel_row, 10).value = effect_system or ""

    # SG (col 11)
    ws.cell(excel_row, 11).value = sg or ""

    # S (col 12)
    ws.cell(excel_row, 12).value = S

    # Preventive Action (col 13)
    ws.cell(excel_row, 13).value = prev_action or ""

    # O (col 14)
    ws.cell(excel_row, 14).value = O

    # Safety Mechanism (col 15)
    ws.cell(excel_row, 15).value = sm or ""

    # Test Method (col 16)
    ws.cell(excel_row, 16).value = test_sm or ""

    # D_sm (col 17)
    ws.cell(excel_row, 17).value = D_sm

    # D_test (col 18)
    ws.cell(excel_row, 18).value = D_test

    # D = MIN(D_sm, D_test) (col 19) — formula or computed
    if D_sm is not None and D_test is not None:
        d_col17 = get_column_letter(17)
        d_col18 = get_column_letter(18)
        ws.cell(excel_row, 19).value = (
            f"=IF(MIN({d_col17}{excel_row}:{d_col18}{excel_row}),"
            f"MIN({d_col17}{excel_row}:{d_col18}{excel_row}),\"\")"
        )
    else:
        ws.cell(excel_row, 19).value = ""

    # RPN = S * O * D (col 20)
    if S is not None and O is not None and D_sm is not None:
        s_col = get_column_letter(12)
        o_col = get_column_letter(14)
        d_col = get_column_letter(19)
        ws.cell(excel_row, 20).value = (
            f"=IFERROR({s_col}{excel_row}*{o_col}{excel_row}*{d_col}{excel_row},0)"
        )
    else:
        ws.cell(excel_row, 20).value = ""

    # Countermeasure (col 21)
    ws.cell(excel_row, 21).value = countermeasure or ""

    # Apply styles to all cols
    for col_j in range(1, 22):
        cell = ws.cell(excel_row, col_j)
        cell.fill = fill
        cell.border = thin_border()
        if col_j in (1, 12, 14, 17, 18, 19, 20, 21):
            cell.alignment = wrap_center()
        else:
            cell.alignment = wrap_left()
        cell.font = NORMAL

    ws.row_dimensions[excel_row].height = 45

# ── Merge variable name / type / sw_unit / category cells ───
last_row_in_sheet = START_ROW + len(rows_data) - 1
# Also update the last variable's end row
if current_var_name is not None:
    var_block_rows[current_var_name][1] = last_row_in_sheet

for var_name_key, (start_r, end_r) in var_block_rows.items():
    if end_r > start_r:
        for col_m in [2, 3, 4, 5]:  # SW Unit, Category, Var name, Var type
            try:
                ws.merge_cells(
                    start_row=start_r, start_column=col_m,
                    end_row=end_r, end_column=col_m
                )
                merged_cell = ws.cell(start_r, col_m)
                merged_cell.alignment = wrap_center()
            except Exception:
                pass

# ── Freeze top rows ──────────────────────────────────────────
ws.freeze_panes = f"A{HDR_ROW+1}"

# ── Save ─────────────────────────────────────────────────────
out_path = "E:/claude/FMEA/CstAp_PwrMGT_FMEA_완성.xlsx"
wb.save(out_path)
print(f"저장 완료: {out_path}")
print(f"총 데이터 행 수: {len(rows_data)}")
content = [r for r in rows_data if r[10] is not None]
na_rows = [r for r in rows_data if r[10] is None]
print(f"  - 분석 행 (S/O/D 있음): {len(content)}")
print(f"  - N/A 행: {len(na_rows)}")
print(f"  - 번호 매긴 항목: {ROW_NO_COUNTER[0]}")
